from urllib import request, error, parse
from bs4 import BeautifulSoup
import openpyxl
import json
import re


class ExcelExporter:
    def __init__(self):
        self.name = ""
        self.websiteAddress = ""
        self.phones = []
        self.emails = []

    @property
    def foundEverything(self):
        if (self.emails or self.phones) and self.name and self.websiteAddress:
            return True
        return False


class NotEnoughInfoError(Exception):
    def __init__(self):
        pass


class MyCrawler:
    links = []

    def __init__(self, result, title):
        try:
            response = request.urlopen(result).read()
        except error.HTTPError:
            return
        soup = BeautifulSoup(response, 'html.parser')  # soup to parse for info
        linksA = soup.find_all('a')

        self.ee = ExcelExporter()  # exporter to populate
        self.ee.websiteAddress = result
        self.ee.name = title

        self.i = 0  # to prevent rabbit holes
        self.startLink = len(MyCrawler.links)
        MyCrawler.links.append(result)
        self.search(result)

        for link in linksA:
            link = link.get('href')
            if link not in MyCrawler.links:
                MyCrawler.links.append(link)
                self.search(link)

        if not self.everythingFound:
            raise NotEnoughInfoError()

    @property
    def everythingFound(self):
        return self.ee.foundEverything

    def findInfo(self, soup):
        text = soup.get_text()
        phoneRegex = re.compile("(\(?\d{3}\)? ?[/\-.]?\d{3}[/\-. ]\d{4})")
        emailRegex = re.compile("([a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+)")
        emails = re.findall(emailRegex, text)
        phoneNumbers = re.findall(phoneRegex, text)
        for email in emails:
            print(email)  #DEBUG
            if email not in self.ee.emails:
                self.ee.emails.append(email)
        for number in phoneNumbers:
            if any(x in number for x in ('410', '443', '667', '202', '571', '703', '540')):
                print(number)  #DEBUG
                self.ee.phones.append(number)

    def findLinks(self, soup):
        newLinks = []
        for a in soup.find_all('a'):
            try:
                link = a.get('href')
                if link not in MyCrawler.links and \
                                self.i < 3 and not \
                                link.endswith(('.jpg', '.png', '.jpeg')) and \
                                link in MyCrawler.links[self.startLink]:
                    MyCrawler.links.append(link)
                    newLinks.append(link)
            except AttributeError:
                continue
        self.i += 1
        return newLinks
        # finds all links on the web page and cross references them with current
        # list to prevent infinite loops, returns all new links

    def search(self, link):
        mainLink = re.compile('http[s]?:\/\/[^\/]+').findall(MyCrawler.links[self.startLink])[0]
        if self.i >= 3:
            return
        else:
            try:
                response = request.urlopen(link).read()
            except ValueError:
                try:

                    for location in link.split('/'):
                        if not location:
                            continue
                        if location in MyCrawler.links[self.startLink]:
                            link = link.replace(location+"/", '')

                    if link == '/':
                        link = ''
                    print(mainLink+'/'+link)
                    response = request.urlopen(mainLink+'/'+link).read()
                except error.HTTPError:
                    return
            except error.URLError:
                return
            soup = BeautifulSoup(response, 'html.parser')

            self.findInfo(soup)
            newLinks = self.findLinks(soup)
            for link in newLinks:
                self.search(link)


def crawl(result, title):
    try:
        crawler = MyCrawler(result, title)
        return crawler.ee
    except NotEnoughInfoError:
        return None


def doSearch():
    wb = openpyxl.Workbook()
    sheet1 = wb.active
    sheet1.title = 'Sub Contacts'
    queries = ('electrical', 'masonry', 'plumbing', 'HVAC', 'surveying', 'concrete')

    i = 1
    for query in queries:
        q = query+" contractor calvert county maryland"
        q = parse.urlencode({"q": q})
        print(q)  # DEBUG
        response = json.loads(request.urlopen("https://www.googleapis.com/customsearch/v1?"+q+
                                                     "&cx=000000000000000000000:00000000000"+
                                                     "&key=0000000000000000000_0000000000"+
                                                     "&excludeTerms=best+top+chamber"+
                                                     "&num=10").read())

        for result in response['items']:
            print(result["link"])  # DEBUG
            try:
                ee = crawl(result['link'], result['title'])
                sheet1.cell(row=i, column=1, value=ee.name)
                sheet1.cell(row=i, column=2, value=ee.websiteAddress)

                for k in range(1, len(ee.emails)):
                    sheet1.cell(row=i+k, column=3, value=ee.emails[k])
                for k in range(1, len(ee.phones)):
                    sheet1.cell(row=i+k, column=4, value=ee.phones[k])

                j = len(ee.phones)
                if len(ee.emails) > j:
                    j = len(ee.emails)
                i = i + j + 1
            except AttributeError:
                continue

    wb.save('Sub Contacts.xlsx')


def main():
    doSearch()


if __name__ == "__main__":
    main()
